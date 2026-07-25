#!/usr/bin/env python3
"""Build historical abbreviation controls for the variable-length decoder.

The output combines two controls with different evidential roles:

* Cappelli is a Latin abbreviation lexicon.  We retain records whose stated
  period intersects the fourteenth or fifteenth century.  Its bracket notation
  is preserved without assuming that bracketed components are omitted letters.
* The Nuremberg Letterbooks are running early-fifteenth-century text.  Their
  diplomatic PAGE XML explicitly marks supplied letters as
  ``<expan>visible<ex>omitted</ex></expan>``.  We extract only TextLine Unicode
  nodes, avoiding the regularized full-page copy also embedded in each file.

Nuremberg parsing uses only the standard library.  Cappelli XLSX parsing
requires openpyxl, but ``--only nuremberg`` does not.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree
from zipfile import ZipFile


HERE = Path(__file__).resolve().parent
ROOT = HERE.parents[1]
DEFAULT_CAPPELLI = Path("/tmp/cappelli_data/cappelli.xlsx")
DEFAULT_NUREMBERG = Path("/tmp/nuremberg_labels.zip")
DEFAULT_OUTPUT = ROOT / "data" / "external" / "historical_abbreviation_controls.json"

CAPPELLI_PROVENANCE = {
    "title": "Cappelli: Downloadable Data",
    "publisher": "University of Zurich, Ad fontes",
    "source_url": (
        "https://www.adfontes.uzh.ch/en/ressourcen/abkuerzungen/"
        "cappelli-daten-zum-download"
    ),
    "original_work": (
        "Adriano Cappelli, Lexicon Abbreviaturarum, second edition, "
        "Leipzig, 1928"
    ),
    "rights": {
        "status": "public_domain",
        "statement": (
            "Ad fontes states that the images and information derive from "
            "Cappelli's Lexicon, are in the public domain, and may be reused "
            "without restriction."
        ),
    },
    "data_note": (
        "Crowdsourced transcription with crowd and expert quality control; "
        "Ad fontes cautions that occasional errors may remain and that its "
        "download is generated from the most recent data."
    ),
}

NUREMBERG_PROVENANCE = {
    "title": (
        "Nuremberg Letterbooks: A Multi-Transcriptional Dataset of Early "
        "15th Century Manuscripts for Document Analysis"
    ),
    "authors": [
        "Martin Mayr",
        "Julian Krenz",
        "Katharina Neumeier",
        "Anna Bub",
        "Simon Bürcky",
        "Nina Brolich",
        "Klaus Herbers",
        "Mechthild Habermann",
        "Peter Fleischmann",
        "Andreas Maier",
        "Vincent Christlein",
    ],
    "version": "v1",
    "publication_date": "2024-10-02",
    "dataset_doi": "10.5281/zenodo.13881575",
    "dataset_url": "https://zenodo.org/records/13881575",
    "article_doi": "10.1038/s41597-025-05144-z",
    "license": {
        "spdx": "CC-BY-4.0",
        "name": "Creative Commons Attribution 4.0 International",
        "url": "https://creativecommons.org/licenses/by/4.0/",
    },
    "coverage": (
        "Nuremberg Letterbooks 2-5, February 7, 1408 through March 31, 1423"
    ),
    "source_file": "labels.zip",
    "zenodo_md5": "ce2c6150d9fc45ac4b4ea2a439b7aa8e",
}

ROMAN_VALUES = {
    "I": 1,
    "V": 5,
    "X": 10,
    "L": 50,
    "C": 100,
    "D": 500,
    "M": 1000,
}
ROMAN_TOKEN = re.compile(r"[IVX]+")
CAPPELLI_MARKED_COMPONENT = re.compile(r"\[([^\]]*)\]")
NUREMBERG_MEMBER = re.compile(
    r"^nuremberg_letterbooks/diplomatic-regularised/Band([2-5])/.+\.xml$"
)
DECODER_TARGET_CHUNKS = (
    "us",
    "um",
    "con",
    "com",
    "cum",
    "per",
    "pro",
    "pre",
    "et",
    "er",
    "que",
    "qui",
)


def normalized(value: Any) -> str:
    """Return stable NFC text while preserving internal source whitespace."""

    if value is None:
        return ""
    return unicodedata.normalize("NFC", str(value).strip())


def normalized_token(value: Any) -> str:
    """Normalize XML formatting whitespace inside one expansion token."""

    return re.sub(r"\s+", " ", normalized(value))


def file_digests(path: Path) -> dict[str, Any]:
    sha256 = hashlib.sha256()
    md5 = hashlib.md5(usedforsecurity=False)
    size = 0
    with path.open("rb") as handle:
        while chunk := handle.read(1024 * 1024):
            size += len(chunk)
            sha256.update(chunk)
            md5.update(chunk)
    return {
        "path_at_build": str(path),
        "size_bytes": size,
        "sha256": sha256.hexdigest(),
        "md5": md5.hexdigest(),
    }


def roman_to_int(token: str) -> int:
    total = 0
    for index, character in enumerate(token):
        value = ROMAN_VALUES[character]
        following = (
            ROMAN_VALUES[token[index + 1]] if index + 1 < len(token) else 0
        )
        total += -value if value < following else value
    return total


def period_intersects_xiv_xv(period: str) -> bool:
    """Select point dates or stated ranges intersecting centuries XIV-XV."""

    values = [roman_to_int(token) for token in ROMAN_TOKEN.findall(period)]
    if any(value in (14, 15) for value in values):
        return True
    if "-" not in period or len(values) < 2:
        return False
    low, high = min(values), max(values)
    return any(low <= target <= high for target in (14, 15))


def parse_unsure(value: Any) -> bool:
    return normalized(value).casefold() not in {"", "0", "false", "none"}


def build_cappelli(path: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise SystemExit(
            "Cappelli extraction requires openpyxl; install it or use "
            "--only nuremberg."
        ) from error

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    header = tuple(next(rows))
    indexes = {name: index for index, name in enumerate(header)}
    required = {
        "id",
        "page_id",
        "characters",
        "transcription",
        "category",
        "period",
        "language",
        "unsure",
    }
    missing = required - indexes.keys()
    if missing:
        raise ValueError(f"Cappelli workbook lacks columns: {sorted(missing)}")

    source_record_count = 0
    selected_record_count = 0
    period_counts: Counter[str] = Counter()
    category_counts: Counter[str] = Counter()
    records: dict[
        tuple[str, str, str | None, str, bool],
        dict[str, Any],
    ] = {}

    for row in rows:
        source_record_count += 1
        language = normalized(row[indexes["language"]]).casefold()
        period = normalized(row[indexes["period"]])
        if language != "latin" or not period_intersects_xiv_xv(period):
            continue

        selected_record_count += 1
        written = normalized(row[indexes["characters"]])
        transcription = normalized(row[indexes["transcription"]])
        category_value = normalized(row[indexes["category"]])
        category = category_value or None
        unsure = parse_unsure(row[indexes["unsure"]])
        source_id = int(row[indexes["id"]])
        page_id = int(row[indexes["page_id"]])
        key = (written, transcription, category, period, unsure)

        if key not in records:
            marked_components = [
                normalized(component)
                for component in CAPPELLI_MARKED_COMPONENT.findall(written)
            ]
            records[key] = {
                "written_notation": written,
                "visible": CAPPELLI_MARKED_COMPONENT.sub("", written),
                "marked_components": marked_components,
                "transcription": transcription,
                "period": period,
                "category": category,
                "unsure": unsure,
                "source_ids": [],
                "source_page_ids": [],
                "count": 0,
            }
        record = records[key]
        record["source_ids"].append(source_id)
        record["source_page_ids"].append(page_id)
        record["count"] += 1
        period_counts[period] += 1
        category_counts[category or ""] += 1

    workbook.close()

    output_records = list(records.values())
    for record in output_records:
        record["source_ids"] = sorted(set(record["source_ids"]))
        record["source_page_ids"] = sorted(set(record["source_page_ids"]))
    output_records.sort(
        key=lambda record: (
            record["written_notation"].casefold(),
            record["transcription"].casefold(),
            record["period"],
            record["category"] or "",
            record["unsure"],
        )
    )

    if sum(record["count"] for record in output_records) != selected_record_count:
        raise AssertionError("Cappelli deduplication changed the selected row count")

    chunk_summary: dict[str, dict[str, Any]] = {}
    for chunk in DECODER_TARGET_CHUNKS:
        matching = [
            record
            for record in output_records
            if chunk in record["transcription"].casefold()
        ]
        chunk_summary[chunk] = {
            "deduplicated_record_count": len(matching),
            "source_row_count": sum(record["count"] for record in matching),
        }

    return {
        "selection": (
            "language equals latin (case-insensitive), and the period point or "
            "range intersects century XIV or XV"
        ),
        "notation_note": (
            "visible removes square-bracketed components from Cappelli's "
            "characters field. marked_components preserves their content. "
            "No claim is made that those components are omitted plaintext."
        ),
        "source_record_count": source_record_count,
        "selected_record_count": selected_record_count,
        "deduplicated_record_count": len(output_records),
        "period_counts": dict(sorted(period_counts.items())),
        "category_counts": dict(
            sorted(category_counts.items(), key=lambda item: (-item[1], item[0]))
        ),
        "licensed_chunk_summary": {
            "method": (
                "Case-insensitive literal substring match within Cappelli's "
                "transcription/solution field. Counts license a chunk as part "
                "of an attested solution, not necessarily as the omitted span."
            ),
            "chunks": chunk_summary,
        },
        "records": output_records,
    }


def local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def descendants(
    element: ElementTree.Element,
    name: str,
) -> Iterable[ElementTree.Element]:
    return (child for child in element.iter() if local_name(child.tag) == name)


def text_without_ex(element: ElementTree.Element) -> str:
    parts: list[str] = [element.text or ""]
    for child in element:
        if local_name(child.tag) != "ex":
            parts.append(text_without_ex(child))
        parts.append(child.tail or "")
    return "".join(parts)


def writer_sort_key(writer_id: str) -> tuple[Any, ...]:
    components = writer_id.split(",")
    if all(component.isdigit() for component in components):
        return (0, *(int(component) for component in components))
    return (1, writer_id)


def build_nuremberg(path: Path) -> dict[str, Any]:
    pair_counts: defaultdict[
        tuple[int, str],
        Counter[tuple[str, str, str]],
    ] = defaultdict(Counter)
    group_files: defaultdict[tuple[int, str], set[str]] = defaultdict(set)
    book_events: Counter[int] = Counter()
    book_files: defaultdict[int, set[str]] = defaultdict(set)
    xml_file_count = 0
    line_count = 0
    tagged_expan_count = 0
    excluded_without_ex = 0
    excluded_empty = 0

    with ZipFile(path) as archive:
        members = [
            (name, match)
            for name in archive.namelist()
            if (match := NUREMBERG_MEMBER.match(name))
        ]
        for member, match in sorted(members):
            book = int(match.group(1))
            root = ElementTree.fromstring(archive.read(member))
            xml_file_count += 1
            book_files[book].add(member)

            for line in descendants(root, "TextLine"):
                line_count += 1
                writer_id = normalized(line.attrib.get("writerID")) or "unknown"
                unicode_element = next(descendants(line, "Unicode"), None)
                if unicode_element is None:
                    continue

                for expan in descendants(unicode_element, "expan"):
                    tagged_expan_count += 1
                    ex_elements = list(descendants(expan, "ex"))
                    if not ex_elements:
                        excluded_without_ex += 1
                        continue

                    omitted = normalized_token(
                        "".join(
                            "".join(ex_element.itertext())
                            for ex_element in ex_elements
                        )
                    )
                    visible = normalized_token(text_without_ex(expan))
                    expanded = normalized_token("".join(expan.itertext()))
                    if not omitted or not expanded:
                        excluded_empty += 1
                        continue

                    group = (book, writer_id)
                    pair_counts[group][(visible, omitted, expanded)] += 1
                    group_files[group].add(member)
                    book_events[book] += 1

    groups: list[dict[str, Any]] = []
    for book, writer_id in sorted(
        pair_counts, key=lambda value: (value[0], writer_sort_key(value[1]))
    ):
        counts = pair_counts[(book, writer_id)]
        pairs = [
            {
                "visible": visible,
                "omitted": omitted,
                "expanded": expanded,
                "count": count,
            }
            for (visible, omitted, expanded), count in sorted(
                counts.items(),
                key=lambda item: (
                    -item[1],
                    item[0][0].casefold(),
                    item[0][1].casefold(),
                    item[0][2].casefold(),
                ),
            )
        ]
        groups.append(
            {
                "book": book,
                "writer_id": writer_id,
                "source_file_count": len(group_files[(book, writer_id)]),
                "event_count": sum(counts.values()),
                "unique_pair_count": len(pairs),
                "pairs": pairs,
            }
        )

    usable_event_count = sum(group["event_count"] for group in groups)
    if usable_event_count != sum(book_events.values()):
        raise AssertionError("Nuremberg group and book event counts disagree")
    if tagged_expan_count != (
        usable_event_count + excluded_without_ex + excluded_empty
    ):
        raise AssertionError("Nuremberg expansion accounting is incomplete")

    return {
        "selection": (
            "PAGE XML members under diplomatic-regularised/Band2-5; expan "
            "elements under line-level TextLine/TextEquiv/Unicode with a "
            "non-empty ex descendant"
        ),
        "text_note": (
            "visible is diplomatic text outside ex; omitted is concatenated "
            "text inside ex; expanded preserves source order across both. "
            "Writer labels are retained verbatim, including composite labels."
        ),
        "xml_file_count": xml_file_count,
        "line_count": line_count,
        "tagged_expan_count": tagged_expan_count,
        "usable_event_count": usable_event_count,
        "excluded_without_ex_count": excluded_without_ex,
        "excluded_empty_count": excluded_empty,
        "unique_grouped_pair_count": sum(
            group["unique_pair_count"] for group in groups
        ),
        "book_counts": {
            str(book): {
                "source_file_count": len(book_files[book]),
                "event_count": book_events[book],
            }
            for book in sorted(book_files)
        },
        "groups": groups,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--cappelli", type=Path, default=DEFAULT_CAPPELLI)
    parser.add_argument("--nuremberg", type=Path, default=DEFAULT_NUREMBERG)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument(
        "--only",
        choices=("all", "cappelli", "nuremberg"),
        default="all",
        help="Build both controls or one independently.",
    )
    parser.add_argument(
        "--pretty",
        action="store_true",
        help="Indent JSON for inspection; the committed artifact is compact.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    include_cappelli = args.only in {"all", "cappelli"}
    include_nuremberg = args.only in {"all", "nuremberg"}
    for path, included in (
        (args.cappelli, include_cappelli),
        (args.nuremberg, include_nuremberg),
    ):
        if included and not path.is_file():
            raise SystemExit(f"Input file not found: {path}")

    payload: dict[str, Any] = {
        "schema_version": 1,
        "purpose": (
            "Historically attested controls for abbreviation-aware, "
            "variable-length Voynich decoding"
        ),
        "provenance": {},
    }
    if include_cappelli:
        cappelli_provenance = dict(CAPPELLI_PROVENANCE)
        cappelli_provenance["input_file"] = file_digests(args.cappelli)
        payload["provenance"]["cappelli"] = cappelli_provenance
        payload["cappelli_latin_xiv_xv"] = build_cappelli(args.cappelli)
    if include_nuremberg:
        nuremberg_provenance = dict(NUREMBERG_PROVENANCE)
        nuremberg_provenance["input_file"] = file_digests(args.nuremberg)
        if (
            nuremberg_provenance["input_file"]["md5"]
            != NUREMBERG_PROVENANCE["zenodo_md5"]
        ):
            raise ValueError("Nuremberg labels.zip does not match the Zenodo MD5")
        payload["provenance"]["nuremberg_letterbooks"] = nuremberg_provenance
        payload["nuremberg_letterbooks"] = build_nuremberg(args.nuremberg)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    if args.pretty:
        rendered = json.dumps(payload, ensure_ascii=False, indent=2)
    else:
        rendered = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    args.output.write_text(rendered + "\n", encoding="utf-8")

    summary = {
        "output": str(args.output),
        "size_bytes": args.output.stat().st_size,
    }
    if include_cappelli:
        summary["cappelli_selected"] = payload["cappelli_latin_xiv_xv"][
            "selected_record_count"
        ]
        summary["cappelli_unique"] = payload["cappelli_latin_xiv_xv"][
            "deduplicated_record_count"
        ]
    if include_nuremberg:
        summary["nuremberg_events"] = payload["nuremberg_letterbooks"][
            "usable_event_count"
        ]
        summary["nuremberg_grouped_pairs"] = payload[
            "nuremberg_letterbooks"
        ]["unique_grouped_pair_count"]
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
